import openpyxl
import csv
import io
from .models import *

wb = openpyxl.load_workbook(r"C:\Users\carlo\Downloads\26.05.23.xlsx", data_only=True)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column

for row in ws.iter_rows(max_row=max_row, max_col=max_col):
    print(row[0].value, row[1].value)

Remessa.objects.all().delete()

def save_data(data):
    '''
    Salva os dados no banco.
    '''
    aux = []
    for item in data:
        title = item.get('title')
        obj = Remessa(
            title=title,
        )
        aux.append(obj)

    Remessa.objects.bulk_create(aux)

data = []

for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
    _dict = dict(title=row[0].value)
    data.append(_dict)

save_data(data)